#!/usr/bin/env python3

from __future__ import annotations

import csv
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from textwrap import dedent

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    fallback_python = Path("/Users/bai/anaconda3/bin/python")
    if Path(sys.executable) != fallback_python and fallback_python.exists():
        os.execv(str(fallback_python), [str(fallback_python), __file__, *sys.argv[1:]])
    raise


VAULT = Path("/Users/bai/Documents/Obsidian/Space")
DOWNLOADS = Path("/Users/bai/Downloads")
PLAYGROUND = Path("/Users/bai/Documents/Playground")

GO_XLSX = next(iter(sorted(DOWNLOADS.glob("GO*.xlsx"))))
DEVONICS_TRANSCRIPT = DOWNLOADS / "cleaned_meeting_transcript_labeled.txt"
WELLWIT_ISSUES = PLAYGROUND / "wellwit_issue_history_clean.csv"
GO_SOURCE_LABEL = "GreyOrange internal workbook"


def quote(value):
    if value is None:
        return '""'
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return json.dumps(str(value), ensure_ascii=False)


def quote_list(values):
    return "[" + ", ".join(quote(v) for v in values) + "]"


def frontmatter(data):
    lines = ["---"]
    for key, value in data.items():
        if isinstance(value, list):
            lines.append(f"{key}: {quote_list(value)}")
        else:
            lines.append(f"{key}: {quote(value)}")
    lines.append("---")
    return "\n".join(lines)


def write_note(path: Path, fm: dict, body: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    content = frontmatter(fm) + "\n\n" + body.strip() + "\n"
    path.write_text(content, encoding="utf-8")


def safe_name(name: str) -> str:
    name = re.sub(r"[\\/:*?\"<>|]", "-", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name.rstrip(".")


def wiki(path: str, label: str | None = None) -> str:
    name = Path(path).stem
    if label:
        return f"[[{path}|{label}]]"
    return f"[[{path}|{name}]]"


TEAM_MAP = {
    "\u4f9b\u5e94\u94fe": "Supply Chain",
    "\u6280\u672f\u56e2\u961f": "Technical",
    "\u8f6f\u4ef6": "Software",
    "\u8d28\u91cf\u56e2\u961f": "Quality",
}


ROLE_HINTS = {
    "Sumit Tiwary": "Supply chain director-level stakeholder",
    "Harshita Mahajan": "Primary supply chain contact",
    "Varun sharma": "Supply chain coordinator",
    "Mohit Kumar": "Technical lead",
    "Akhil Kumar Verma": "Solutions / program contact",
    "Rishav Kumar": "Solutions / program contact",
    "Nadi": "Solutions / program contact",
    "Vishal Singh": "Field operations lead",
    "Sahil bansal": "Field operations contact",
    "Jale": "Software lead",
    "San": "Quality lead",
    "Kaushal": "Quality contact",
    "Rahul": "Quality / field contact",
}


GREYORANGE_PUBLIC = {
    "website": "https://www.greyorange.com/",
    "linkedin": "https://www.linkedin.com/company/greyorange/",
    "notes": [
        "Workbook source: GreyOrange internal workbook.",
        "Company website and LinkedIn company page checked manually on 2026-04-05.",
    ],
}


WELLWIT_PUBLIC = {
    "website": "https://wellwit.com/",
    "linkedin": "",
    "notes": [
        "Public company site checked manually on 2026-04-05.",
        "Product family reference: https://wellwit.com/fmr-series-2/",
    ],
}


SITE_NORMALIZATION = {
    "adams": "Adams",
    "adam": "Adams",
    "dfw": "DFW",
    "gmi social": "GMI Social",
    "h&m canada": "H&M",
    "india lab": "India Lab",
    "kenco": "Kenco",
    "kenco kansas": "Kenco Kansas",
    "kenco goodyear": "Kenco Goodyear",
    "dorman warsaw": "Dorman Warsaw",
    "dorman portland": "Dorman Portland",
    "sodimac colombia": "Sodimac Colombia",
    "sodimac columbia": "Sodimac Colombia",
    "sodimac chile": "Sodimac Chile",
    "gfc-2": "GFC",
    "gfc": "GFC",
    "sam": "Sam's ATL",
    "sam's atl": "Sam's ATL",
    "sams atl": "Sam's ATL",
    "ykkap": "YKK",
    "ykk": "YKK",
    "rialto": "Kellanova Rialto",
    "envista": "Kellanova Envista",
    "minooka": "Kellanova Minooka",
    "greenbox": "Greenbox",
    "casey": "Casey",
    "fairlife": "Fairlife",
    "kerry": "Kerry",
    "reckitt": "Kenco Reckitt",
    "walmart hauling solution": "Walmart Hauling Solution",
    "leadership level": "Unspecified",
    "stp": "Unspecified",
    "the bot level": "Unspecified",
    "the earliest": "Unspecified",
    "the end of march": "Unspecified",
    "the factory": "Unspecified",
    "the pre-lm point": "Unspecified",
    "your end": "Unspecified",
}


DEVONICS_CONTACTS = [
    {
        "name": "Joe Fautas",
        "organization": "Devonics",
        "account_path": "09 Work/Accounts/Devonics.md",
        "team": "Commercial",
        "title": "Co-founder; sales and marketing",
        "relationship": "active",
        "verification_status": "transcript_derived",
        "verification_sources": [str(DEVONICS_TRANSCRIPT)],
        "notes": [
            "Introduced himself as running sales and marketing.",
            "Said he is a co-founder of Fairino USA and Devonix Automation.",
        ],
    },
    {
        "name": "Frank Handel",
        "organization": "Devonics",
        "account_path": "09 Work/Accounts/Devonics.md",
        "team": "Engineering",
        "title": "CTO; VP Engineering",
        "relationship": "active",
        "verification_status": "transcript_derived",
        "verification_sources": [str(DEVONICS_TRANSCRIPT)],
        "notes": [
            "Joined late in the call and introduced himself as CTO and VP of engineering.",
            "Owns optimization, customization, and support topics.",
        ],
    },
    {
        "name": "Joy",
        "organization": "Devonics",
        "account_path": "09 Work/Accounts/Devonics.md",
        "team": "Supplier Coordination",
        "title": "China-based supplier and project coordination",
        "relationship": "active",
        "verification_status": "transcript_derived",
        "verification_sources": [str(DEVONICS_TRANSCRIPT)],
        "notes": [
            "Introduced as the China-side team member coordinating suppliers and projects.",
        ],
    },
    {
        "name": "Darshan",
        "organization": "Outland Robotics",
        "account_path": "09 Work/Accounts/Outland Robotics.md",
        "team": "Founding Team",
        "title": "Co-founder",
        "relationship": "active",
        "verification_status": "transcript_derived",
        "verification_sources": [str(DEVONICS_TRANSCRIPT)],
        "notes": [
            "Presented Outland Robotics use cases and requirements for a mobile manipulation platform.",
        ],
    },
    {
        "name": "Leo",
        "organization": "Outland Robotics",
        "account_path": "09 Work/Accounts/Outland Robotics.md",
        "team": "Founding Team",
        "title": "CEO; CTO",
        "relationship": "active",
        "verification_status": "transcript_derived",
        "verification_sources": [str(DEVONICS_TRANSCRIPT)],
        "notes": [
            "Introduced as CEO and CTO of Outland Robotics.",
            "Asked about deployments, applications, and commercialization path.",
        ],
    },
    {
        "name": "Samantha",
        "organization": "Wellwit Robotics",
        "account_path": "09 Work/Accounts/Wellwit Robotics.md",
        "team": "Sales",
        "title": "Overseas sales manager",
        "relationship": "active",
        "verification_status": "company_verified_transcript_contact",
        "verification_sources": [str(DEVONICS_TRANSCRIPT), WELLWIT_PUBLIC["website"]],
        "notes": [
            "Introduced as overseas sales manager of Wellwit and owner of the Devonics project.",
        ],
    },
    {
        "name": "Vincent",
        "organization": "Wellwit Robotics",
        "account_path": "09 Work/Accounts/Wellwit Robotics.md",
        "team": "Software",
        "title": "Software engineer",
        "relationship": "active",
        "verification_status": "company_verified_transcript_contact",
        "verification_sources": [str(DEVONICS_TRANSCRIPT), WELLWIT_PUBLIC["website"]],
        "notes": [
            "Atlanta-based software engineer from the Wellwit team.",
            "Covered power budget, SLAM, maintenance, and integration details.",
        ],
    },
]


def normalize_site(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return "Unspecified"
    key = text.lower()
    if key in SITE_NORMALIZATION:
        return SITE_NORMALIZATION[key]
    for pattern, normalized in SITE_NORMALIZATION.items():
        if pattern in key:
            return normalized
    return re.sub(r"\s+", " ", text.title())


def issue_is_useful(row: dict) -> bool:
    site = normalize_site(row.get("site_std") or row.get("site"))
    category = (row.get("category_std") or row.get("category") or "").strip()
    problem = (row.get("problem_description_clean") or row.get("problem_description") or "").strip()
    title = (row.get("title_clean") or row.get("title") or "").strip()
    if site != "Unspecified":
        return True
    if category in {
        "hardware_or_mechanical",
        "software_or_firmware",
        "product_design",
        "site_support",
        "structured_bug_list",
    }:
        return True
    if len(problem) >= 120 or len(title) >= 80:
        return True
    return False


def short_issue_title(row: dict) -> str:
    title = (row.get("issue_summary") or row.get("title_clean") or row.get("title") or "Issue").strip()
    title = re.sub(r"\s+", " ", title)
    if len(title) > 90:
        title = title[:87].rstrip() + "..."
    return title


def normalize_date(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return "undated"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%y, %I:%M %p",
        "%m/%d/%Y, %I:%M %p",
        "%m/%d/%y",
        "%m/%d/%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if match:
        return match.group(0)
    return "undated"


def translate_go_contact_note(note: str) -> str:
    text = re.sub(r"\s+", " ", (note or "")).strip()
    if not text:
        return "Role detail still needs confirmation from live collaboration."
    replacements = {
        "\u7c7b\u4f3c\u65b9\u6848\u7ecf\u7406\u4e4b\u7c7b\u7684": "Acts like a solutions or program manager.",
        "\u65b9\u6848\u7ecf\u7406\u4e4b\u7c7b\u7684": "Acts like a solutions or program manager.",
        "\u8d1f\u8d23\u73b0\u573a\u7684\u4eba": "Owns field operations for active sites.",
        "Anand\u7684\u4e0b\u5c5e": "Direct report to Anand.",
        "\u4f9b\u5e94\u94fe\u4e3b\u8981\u5bf9\u63a5\u4eba\uff0charshita \u7684\u4e0b\u5c5e": "Supply chain contact working under Harshita.",
        "\u4e3b\u8981\u5bf9\u63a5\u56e2\u961f\uff0c\u8d28\u91cf\u8d1f\u8d23\u4eba\uff0c\u8d28\u91cf\u7684\u8001\u5927": "Primary quality stakeholder and leader for the quality team.",
        "\u4f9b\u5e94\u94fe\u4e3b\u8981\u5bf9\u63a5\u4eba\uff0c": "Primary supply chain contact.",
        "\u8d28\u91cf\u8d1f\u8d23\u4eba": "Quality lead.",
        "\u8f6f\u4ef6\u8d1f\u8d23\u4eba": "Software lead.",
        "\u603b\u76d1\u7ea7\u522b\u4eba\u7269\uff0c\u4e00\u822c\u4e0d\u53d1\u90ae\u4ef6\u7ed9\u4ed6": "Director-level stakeholder; usually not the default email contact.",
        "\u6280\u672f\u8d1f\u8d23\u4eba": "Technical lead.",
    }
    return replacements.get(text, text)


def normalize_fleet_text(text: str) -> str:
    value = re.sub(r"\s+", " ", (text or "")).strip()
    if not value:
        return ""
    value = re.sub(r"(\d+)\s*\u53f0\u53c9\u8f66", r"\1 forklifts", value)
    value = re.sub(r"(\d+)\s*\u53f0\u8f66", r"\1 vehicles", value)
    value = re.sub(r"(\d+)\s*\u53f0\s*([A-Za-z]+)", r"\1 \2 units", value)
    return value


def translate_go_operational_note(note: str) -> str:
    text = re.sub(r"\s+", " ", (note or "")).strip()
    if not text:
        return ""
    replacements = {
        "\u73b0\u573a\u7535\u6c60\u5f02\u5e38\uff0c\u76ee\u524d\u5df2\u4f7f\u7528\u5907\u54c1\u6b63\u5e38\u8fd0\u884c\uff0c\u9700\u8981\u5206\u6790\u62a5\u544a": "Battery anomaly on site. A spare unit is currently keeping operations running. Analysis report still needed.",
        "\u5df2\u505c\u6b62": "Deployment appears stopped.",
        "\u975e\u5e38\u65e7\u7684\u73b0\u573a\uff0c\u4e00\u5171\u670963\u53f0\u8f66": "Very old deployment site with 63 robots in total.",
        "2\u53f0\u53c9\u8f66\uff0c\u9502\u5e73": "Fleet note: 2 forklifts with lithium pallet configuration.",
        "2\u53f0AP\uff0c2\u53f0L \u8f66\uff0c2\u53f0H": "Fleet note: 2 AP units, 2 L units, and 2 H units.",
        "2\u53f0AP\uff0c2\u53f0L \u8f66\uff0c2\u53f0H ": "Fleet note: 2 AP units, 2 L units, and 2 H units.",
    }
    return replacements.get(text, text)


def sanitize_imported_text(text: str) -> str:
    value = (text or "").strip()
    if not value:
        return ""
    replacements = {
        "\u67e5\u8be2\u673a\u5668\u4eba\u4e0d\u6267\u884c\u81ea\u52a8\u4efb\u52a1\u539f\u56e0 - \u98de\u4e66\u4e91\u6587\u6863": "Check why the robot is not executing auto tasks - Feishu doc",
        "\u4e0a\u4f20\u573a\u666f - \u98de\u4e66\u4e91\u6587\u6863": "Upload scene - Feishu doc",
        "CBD30H-YD\u81ea\u52a8\u5bfc\u5f15\u6258\u76d8\u642c\u8fd0\u8f66450\u9ad8\u5bf9\u6bd4\u56fe(1).dwg": "CBD30H-YD automated guided pallet truck 450-lift comparison drawing (1).dwg",
        "CBD30H-YD\u81ea\u52a8\u5bfc\u5f15\u6258\u76d8\u642c\u8fd0\u8f66450\u9ad8.dwg": "CBD30H-YD automated guided pallet truck 450-lift drawing.dwg",
        "\u81ea\u52a8\u5bfc\u5f15\u6258\u76d8\u642c\u8fd0\u8f66": "automated guided pallet truck",
        "greyorange side": "GreyOrange side",
        "\u98de\u4e66\u4e91\u6587\u6863": "Feishu doc",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value


def build_account_note(
    title: str,
    *,
    industry: str,
    region: str,
    website: str = "",
    linkedin: str = "",
    stage: str = "active",
    summary: str = "",
    context: str = "",
    risks: list[str] | None = None,
    next_moves: list[str] | None = None,
) -> tuple[dict, str]:
    risks_block = "\n".join(f"- {item}" for item in (risks or [])) or "-"
    next_block = "\n".join(f"- [ ] {item}" for item in (next_moves or [])) or "- [ ]"
    body = (
        f"# {title}\n\n"
        f"## Account Overview\n\n"
        f"{summary}\n\n"
        f"## Business Context\n\n"
        f"{context}\n\n"
        f"## Stakeholder Map\n\n"
        f"```dataview\n"
        f"TABLE team AS Team, title AS Role, relationship AS Relationship, verification_status AS Verification\n"
        f'FROM "09 Work/Accounts/{title}"\n'
        f'WHERE type = "contact"\n'
        f"SORT team ASC, file.name ASC\n"
        f"```\n\n"
        f"## Active Deployments / Projects\n\n"
        f"```dataview\n"
        f"TABLE project_kind AS Type, site AS Site, status AS Status, next AS Next\n"
        f'FROM "04 Projects/Work"\n'
        f'WHERE type = "project" AND account = "{title}"\n'
        f"SORT file.name ASC\n"
        f"```\n\n"
        f"## Recent Meetings\n\n"
        f"```dataview\n"
        f"TABLE date AS Date, organizations AS Organizations, project AS Project\n"
        f'FROM "09 Work/Meetings"\n'
        f'WHERE type = "meeting" AND contains(organizations, "{title}")\n'
        f"SORT date DESC\n"
        f"```\n\n"
        f"## Open Issues\n\n"
        f"```dataview\n"
        f"TABLE site AS Site, category AS Category, status AS Status, severity AS Severity, last_updated_date AS Updated\n"
        f'FROM "09 Work/Issues"\n'
        f'WHERE type = "issue" AND client = "{title}" AND status != "resolved"\n'
        f"SORT last_updated_date DESC\n"
        f"```\n\n"
        f"## Risks\n\n"
        f"{risks_block}\n\n"
        f"## Next Moves\n\n"
        f"{next_block}"
    )
    fm = {
        "type": "account",
        "account_name": title,
        "stage": stage,
        "industry": industry,
        "region": region,
        "website": website,
        "linkedin": linkedin,
        "owner": "bai",
        "last_contact": "",
        "next_contact": "",
        "tags": [],
    }
    return fm, body


def import_greyorange():
    wb = load_workbook(GO_XLSX, data_only=True)
    people = wb.worksheets[0]
    addresses = wb.worksheets[1]

    account_fm, account_body = build_account_note(
        "GreyOrange",
        industry="Warehouse automation and robotics",
        region="North America; LATAM; India",
        website=GREYORANGE_PUBLIC["website"],
        linkedin=GREYORANGE_PUBLIC["linkedin"],
        summary="GreyOrange should be treated as one strategic account with many deployment sites, many internal stakeholders, and a large support surface.",
        context="The workbook provided is effectively a lightweight account book: it has GreyOrange internal contacts plus a deployment list covering GFC, Adams, Sodimac, Kenco, Kellanova, Sam's ATL, YKK and more. This is exactly the pattern that needs an account-centric system instead of one flat customer note.",
        risks=[
            "Support load is fragmented across many sites and site conditions.",
            "Stakeholders span supply chain, technical, software, field operations, and quality.",
            "Several deployments appear stopped or aging, so historical context matters.",
        ],
        next_moves=[
            "Attach new site issues to the correct deployment note instead of leaving them floating in the inbox.",
            "Promote repeated patterns into SOPs or formal knowledge notes.",
            "Review high-frequency unresolved issues every week and prioritize the sites that keep repeating.",
        ],
    )
    write_note(VAULT / "09 Work/Accounts/GreyOrange.md", account_fm, account_body)

    current_team = ""
    contact_count = 0
    for row in people.iter_rows(min_row=3, values_only=True):
        group, name, email, note = row[1], row[2], row[3], row[4]
        if group:
            current_team = TEAM_MAP.get(group, group)
        if not name:
            continue
        contact_count += 1
        title = ROLE_HINTS.get(name, "GreyOrange stakeholder")
        relationship = "active"
        verification_sources = [GO_SOURCE_LABEL, GREYORANGE_PUBLIC["website"], GREYORANGE_PUBLIC["linkedin"]]
        role_note = translate_go_contact_note(note)
        fm = {
            "type": "contact",
            "organization": "GreyOrange",
            "account": "GreyOrange",
            "team": current_team or "Unknown",
            "title": title,
            "email": email or "",
            "phone": "",
            "location": "",
            "relationship": relationship,
            "owner": "bai",
            "verification_status": "company_verified_internal_contact",
            "verification_sources": verification_sources,
            "last_contact": "",
            "tags": [],
        }
        body = dedent(
            f"""
            # {name}

            ## Profile

            - Organization: GreyOrange
            - Team: {current_team or "Unknown"}
            - Title: {title}
            - Email: {email or ""}
            - Source: {GO_SOURCE_LABEL}

            ## What They Own

            - {role_note}

            ## Relationship Notes

            - This contact came from the GreyOrange workbook maintained for active working relationships.
            - Person-level public profile verification was attempted on 2026-04-05. Where no confident public match was found, the note remains anchored to internal source plus verified company presence.

            ## Verification

            - status: company_verified_internal_contact
            - sources:
              - {GO_SOURCE_LABEL}
              - {GREYORANGE_PUBLIC["website"]}
              - {GREYORANGE_PUBLIC["linkedin"]}

            ## Related Projects / Sites

            ```dataview
            TABLE site, status, next
            FROM "04 Projects/Work"
            WHERE type = "project" AND account = "GreyOrange"
            SORT file.name ASC
            ```
            """
        ).strip()
        write_note(VAULT / f"09 Work/Accounts/GreyOrange/Contacts/{safe_name(name)}.md", fm, body)

    site_count = 0
    for row in addresses.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        site_count += 1
        site = normalize_site(str(row[1]).replace("Columbia", "Colombia").replace("  ", " ").strip())
        fleet = normalize_fleet_text(str(row[3])) if row[3] is not None else ""
        address = row[4] or ""
        raw_note = str(row[5] or "").strip()
        note = translate_go_operational_note(raw_note)
        status = "stopped" if "\u505c\u6b62" in raw_note else "active"
        next_step = ""
        if "\u7535\u6c60\u5f02\u5e38" in raw_note:
            next_step = "Add the battery anomaly analysis conclusion and link the related issues, meetings, and owners."
        elif "\u505c\u6b62" in raw_note:
            next_step = "Confirm whether this site is decommissioned and capture the historical context if it is."
        else:
            next_step = "Link the historical site issues here and update the latest operating status, owner, and next follow-up date."
        fm = {
            "type": "project",
            "account": "GreyOrange",
            "client": "GreyOrange",
            "project_kind": "site_deployment",
            "site": site,
            "status": status,
            "owner": "bai",
            "started": "",
            "next": next_step,
            "tags": [],
        }
        body = (
            f"# GreyOrange - {site}\n\n"
            f"## Goal\n\n"
            f"Track one live or historical GreyOrange deployment site as an operational object that can collect meetings, issues, local context, and follow-ups.\n\n"
            f"## Background\n\n"
            f"- Source workbook: {GO_SOURCE_LABEL}\n"
            f"- Platform / robot type: {row[2] or ''}\n"
            f"- Fleet size: {fleet}\n"
            f"- Notes: {note or ''}\n\n"
            f"## Site / Deployment Information\n\n"
            f"- Customer: GreyOrange\n"
            f"- Site: {site}\n"
            f"- Deployment Type: {row[2] or ''}\n"
            f"- Fleet Size: {fleet}\n"
            f"- Address: {address}\n\n"
            f"## Current Status\n\n"
            f"- Status: {status}\n"
            f'- Operational note: {note or "No additional note provided in workbook."}\n\n'
            f"## Next Steps\n\n"
            f"- [ ] {next_step}\n\n"
            f"## Open Issues\n\n"
            f"```dataview\n"
            f"TABLE category, status, severity, last_updated_date\n"
            f'FROM "09 Work/Issues/GreyOrange"\n'
            f'WHERE type = "issue" AND site = "{site}"\n'
            f"SORT last_updated_date DESC\n"
            f"```"
        )
        write_note(VAULT / f"04 Projects/Work/GreyOrange/{safe_name(site)}.md", fm, body)

    return {"contacts": contact_count, "sites": site_count}


def import_devonics():
    devonics_fm, devonics_body = build_account_note(
        "Devonics",
        industry="Industrial automation integration",
        region="US; China coordination",
        summary="Devonics appears to be the integrator and commercial bridge in this opportunity, connecting Fairino, Wellwit, and Outland Robotics around a mobile manipulation offering.",
        context="The transcript describes a discovery conversation around using a Wellwit mobile base plus a Fairino FR10 and Outland's compute/vision/autonomy stack. Devonics owns customer-facing technical scoping and customization.",
        risks=[
            "Public company profile was not confidently verified from official web sources during this pass.",
            "Opportunity is still in evaluation, so requirements and product boundaries can shift.",
            "Power budget and packaging constraints are open engineering questions.",
        ],
        next_moves=[
            "Capture the proposed architecture as one project note and keep action items linked there.",
            "Confirm whether the external name is Devonics or Devonix Automation in signed documents.",
            "Track package-space, battery, and API integration follow-ups in one opportunity thread.",
        ],
    )
    write_note(VAULT / "09 Work/Accounts/Devonics.md", devonics_fm, devonics_body)

    outland_fm, outland_body = build_account_note(
        "Outland Robotics",
        industry="Robotics integration and autonomy",
        region="US",
        summary="Outland Robotics is the potential downstream product owner or go-to-market partner for a mobile manipulator offering built on Fairino + Wellwit hardware.",
        context="From the transcript, Outland has experience in manipulation, vision, and autonomy, and wants a mobile platform to broaden its offering.",
        risks=[
            "Requirements are exploratory and still market-facing rather than locked to one SKU.",
            "Commercial model and volume assumptions remain early.",
        ],
        next_moves=[
            "Capture target use cases and payload/compute envelope in the linked project note.",
            "Keep all technical trade-offs visible in one place so product assumptions do not drift.",
        ],
    )
    write_note(VAULT / "09 Work/Accounts/Outland Robotics.md", outland_fm, outland_body)

    wellwit_fm, wellwit_body = build_account_note(
        "Wellwit Robotics",
        industry="AMR / AGV platform vendor",
        region="Global; Atlanta support noted in transcript",
        website=WELLWIT_PUBLIC["website"],
        summary="Wellwit is the mobile base vendor in this ecosystem and is already involved in technical feasibility around the 300J platform, power budget, maintenance, and SLAM/API topics.",
        context="This account should be treated as both a vendor relationship and an engineering knowledge source because its platform details directly affect project feasibility and field support.",
        risks=[
            "ROS 2 is not native according to the transcript; integration may require wrappers.",
            "Battery life can be constrained by the proposed GPU-heavy compute stack.",
        ],
        next_moves=[
            "Track platform documentation, maintenance requirements, and integration limits here.",
            "Link future field issues to both customer site and Wellwit platform context.",
        ],
    )
    write_note(VAULT / "09 Work/Accounts/Wellwit Robotics.md", wellwit_fm, wellwit_body)

    for contact in DEVONICS_CONTACTS:
        org = contact["organization"]
        note_lines = "\n".join(f"- {line}" for line in contact["notes"]) or "- Needs follow-up"
        sources_block = "\n".join(f"  - {source}" for source in contact["verification_sources"])
        body = (
            f"# {contact['name']}\n\n"
            f"## Profile\n\n"
            f"- Organization: {org}\n"
            f"- Team: {contact['team']}\n"
            f"- Title: {contact['title']}\n"
            f"- Email:\n"
            f"- Phone:\n"
            f"- Location:\n"
            f"- Source: {DEVONICS_TRANSCRIPT}\n\n"
            f"## What They Own\n\n"
            f"{note_lines}\n\n"
            f"## Relationship Notes\n\n"
            f"- Contact was extracted from the 2026-04-05 Devonics / Outland / Wellwit discovery call.\n"
            f"- Public verification confidence is limited unless a company site was also confirmed.\n\n"
            f"## Verification\n\n"
            f"- status: {contact['verification_status']}\n"
            f"- sources:\n"
            f"{sources_block}\n\n"
            f"## Related Projects / Sites\n\n"
            f"```dataview\n"
            f"TABLE site, status, next\n"
            f'FROM "04 Projects/Work"\n'
            f'WHERE type = "project" AND account = "{org}"\n'
            f"SORT file.name ASC\n"
            f"```"
        )
        fm = {
            "type": "contact",
            "organization": org,
            "account": org,
            "team": contact["team"],
            "title": contact["title"],
            "email": "",
            "phone": "",
            "location": "",
            "relationship": contact["relationship"],
            "owner": "bai",
            "verification_status": contact["verification_status"],
            "verification_sources": contact["verification_sources"],
            "last_contact": "2026-04-05",
            "tags": [],
        }
        write_note(
            VAULT / f"09 Work/Accounts/{org}/Contacts/{safe_name(contact['name'])}.md",
            fm,
            body,
        )

    project_fm = {
        "type": "project",
        "account": "Devonics",
        "client": "Devonics",
        "project_kind": "solution_opportunity",
        "site": "",
        "status": "discovery",
        "owner": "bai",
        "started": "2026-04-05",
        "next": "Collect CAD / internal layout drawings and evaluate power + packaging trade-offs for FR10 on 300J.",
        "tags": [],
    }
    project_body = dedent(
        f"""
        # Devonics - Outland Mobile Manipulation Platform

        ## Goal

        Evaluate a productized mobile manipulator based on a Wellwit mobile base, Fairino FR10 arm, and Outland's autonomy / compute stack.

        ## Background

        - Source transcript: {DEVONICS_TRANSCRIPT}
        - Core organizations: [[09 Work/Accounts/Devonics|Devonics]], [[09 Work/Accounts/Outland Robotics|Outland Robotics]], [[09 Work/Accounts/Wellwit Robotics|Wellwit Robotics]]
        - Candidate base discussed: 300J
        - Arm discussed: FR10

        ## Current Status

        - Stage: discovery
        - Strong fit for warehouse movement and flexible manipulation scenarios.
        - Open questions remain around compute packaging, battery life, ROS 2 integration, and what should be standard vs semi-custom.

        ## Technical Constraints

        - FR10 on Wellwit base appears feasible and close to standard according to Devonics.
        - GPU-heavy compute stack around 500W creates battery pressure.
        - Secondary battery or option A/B/C packaging paths were discussed.
        - Wellwit stated ROS 2 is not native; a wrapper may be required.
        - Weekly maintenance includes LiDAR cleaning and battery-related checks.
        - Clean-room or washdown variants may be possible but are not default.

        ## Next Steps

        - [ ] Receive files / drawings for the three units Darshan showed.
        - [ ] Evaluate 300J internal space against Outland compute box dimensions.
        - [ ] Model battery impact and whether secondary battery is needed.
        - [ ] Decide what belongs in a standard package versus semi-custom upgrade path.
        """
    ).strip()
    write_note(VAULT / "04 Projects/Work/Devonics/Outland Mobile Manipulation Platform.md", project_fm, project_body)

    meeting_fm = {
        "type": "meeting",
        "date": "2026-04-05",
        "organizations": ["Devonics", "Outland Robotics", "Wellwit Robotics"],
        "participants": [c["name"] for c in DEVONICS_CONTACTS],
        "contacts": [c["name"] for c in DEVONICS_CONTACTS],
        "account": "Devonics",
        "project": "Devonics - Outland Mobile Manipulation Platform",
        "client": "Devonics",
        "tags": [],
    }
    meeting_body = dedent(
        f"""
        # 2026-04-05 Devonics / Outland / Wellwit Discovery

        ## Agenda

        - Introductions across Devonics, Outland Robotics, and Wellwit Robotics
        - Evaluate a mobile manipulation offering
        - Discuss deployment fit, platform choice, power budget, software stack, and next technical artifacts

        ## Key Information

        - Outland wants a wider robot offering and is evaluating the right mobile platform.
        - Devonics reported strongest success in warehouse movement use cases, including Walmart and Sam's Club contexts.
        - FR10 on a Wellwit base was described as relatively straightforward.
        - 300J became the candidate platform for the current concept.
        - Outland's compute package is large and power-hungry; around 500W worst-case was discussed.
        - Battery life is therefore a design constraint, with secondary battery and hot-swap options discussed.
        - Wellwit uses LiDAR-based SLAM and said ROS 2 is not native.
        - Weekly maintenance should include LiDAR cleaning and battery checks.
        - Clean-room and IP65-related extensions exist, but are not standard launch assumptions.

        ## Decisions

        - Keep evaluating the 300J for the current project.
        - Treat packaging and power as first-order constraints before committing to a standard product shape.

        ## Action Items

        - [ ] Wellwit / Vincent to provide files or drawings for the three units Darshan showed.
        - [ ] Compare Outland compute dimensions with 300J internal envelope.
        - [ ] Model option A / B / C around battery and accessory battery design.
        - [ ] Clarify API / wrapper approach for ROS-based stack.

        ## Raw Transcript

        - Source file: {DEVONICS_TRANSCRIPT}
        """
    ).strip()
    write_note(
        VAULT / "09 Work/Meetings/Devonics/2026-04-05 Devonics - Outland - Wellwit Discovery.md",
        meeting_fm,
        meeting_body,
    )

    return {"contacts": len(DEVONICS_CONTACTS), "meetings": 1, "projects": 1}


def import_wellwit_issues():
    rows = list(csv.DictReader(WELLWIT_ISSUES.open(encoding="utf-8")))
    imported = 0
    skipped = 0
    per_site = Counter()
    generated_sites = set()
    for index, row in enumerate(rows, start=1):
        if not issue_is_useful(row):
            skipped += 1
            continue
        site = normalize_site(row.get("site_std") or row.get("site"))
        title = sanitize_imported_text(short_issue_title(row))
        created = normalize_date(row.get("created_date") or row.get("created_iso_std") or row.get("created_iso"))
        category = row.get("category_std") or row.get("category") or ""
        status = row.get("status_std") or row.get("status") or ""
        severity = row.get("severity_std") or row.get("severity") or ""
        priority = row.get("priority_std") or ""
        item_type = row.get("item_type_std") or row.get("item_type") or ""
        problem = sanitize_imported_text(row.get("problem_description_clean") or row.get("problem_description") or "")
        progress = sanitize_imported_text(row.get("solution_or_progress_clean") or row.get("solution_or_progress") or "")
        next_action = sanitize_imported_text(row.get("next_action_clean") or row.get("next_action") or "")
        bot_ids = [v.strip() for v in (row.get("bot_ids") or "").split(",") if v.strip()]
        components = [v.strip() for v in (row.get("component_terms") or "").split(",") if v.strip()]
        source_channel = sanitize_imported_text(row.get("source_file") or row.get("source") or "")
        source_id = row.get("source_id") or f"row-{index}"
        imported += 1
        per_site[site] += 1
        if site != "Unspecified":
            generated_sites.add(site)

        fm = {
            "type": "issue",
            "source_type": "support_history_import",
            "client": "GreyOrange",
            "account": "GreyOrange",
            "project": site if site != "Unspecified" else "",
            "site": site,
            "vendor": "Wellwit Robotics",
            "category": category,
            "item_type": item_type,
            "status": status,
            "severity": severity,
            "priority": priority,
            "created_date": created,
            "last_updated_date": row.get("last_updated_date") or "",
            "reporter": row.get("reporter") or "",
            "assignee": row.get("assignee") or "",
            "source_id": source_id,
            "bot_ids": bot_ids,
            "components": components,
            "tags": [],
        }
        body = dedent(
            f"""
            # {title}

            ## Source Context

            - Imported from: {WELLWIT_ISSUES}
            - Source channel/file: {source_channel}
            - Original source id: {row.get("source_id") or ""}
            - Site: {site}
            - Vendor / platform context: Wellwit Robotics

            ## Problem

            {problem or "No normalized problem description available."}

            ## Progress

            {progress or "No progress note captured in the cleaned source."}

            ## Next Action

            {next_action or "No explicit next action found in the cleaned source."}

            ## Metadata

            - Category: {category}
            - Item type: {item_type}
            - Status: {status}
            - Severity: {severity}
            - Priority: {priority}
            - Created: {created}
            - Last updated: {row.get("last_updated_date") or ""}
            - Bot IDs: {", ".join(bot_ids)}
            - Components: {", ".join(components)}

            ## Related Links

            - Account: [[09 Work/Accounts/GreyOrange|GreyOrange]]
            - Site note: {wiki(f"04 Projects/Work/GreyOrange/{safe_name(site)}.md") if site != "Unspecified" else ""}
            - Vendor note: [[09 Work/Accounts/Wellwit Robotics|Wellwit Robotics]]
            """
        ).strip()
        issue_path = VAULT / f"09 Work/Issues/GreyOrange/{safe_name(site)}/{created or 'undated'} - {safe_name(title)} - {safe_name(source_id)}.md"
        write_note(issue_path, fm, body)

    for site in sorted(generated_sites):
        project_path = VAULT / f"04 Projects/Work/GreyOrange/{safe_name(site)}.md"
        if project_path.exists():
            continue
        placeholder_fm = {
            "type": "project",
            "account": "GreyOrange",
            "client": "GreyOrange",
            "project_kind": "site_issue_bucket",
            "site": site,
            "status": "support_only",
            "owner": "bai",
            "started": "",
            "next": "Backfill deployment details when a confirmed source appears.",
            "tags": [],
        }
        placeholder_body = (
            f"# GreyOrange - {site}\n\n"
            f"## Goal\n\n"
            f"Hold issue history for a site or sub-site that appears in support data but was not found in the GreyOrange deployment workbook.\n\n"
            f"## Background\n\n"
            f"- Created automatically from imported issue history.\n"
            f"- Deployment details still need confirmation from live ops records or customer context.\n\n"
            f"## Current Status\n\n"
            f"- Status: support_only\n\n"
            f"## Open Issues\n\n"
            f"```dataview\n"
            f"TABLE category, status, severity, last_updated_date\n"
            f'FROM "09 Work/Issues/GreyOrange"\n'
            f'WHERE type = "issue" AND site = "{site}"\n'
            f"SORT last_updated_date DESC\n"
            f"```"
        )
        write_note(project_path, placeholder_fm, placeholder_body)

    summary_fm = {
        "type": "knowledge",
        "source_type": "import_summary",
        "date": "2026-04-05",
        "tags": [],
    }
    summary_body = (
        f"# Wellwit Issue Import Summary\n\n"
        f"## Import Result\n\n"
        f"- Source file: {WELLWIT_ISSUES}\n"
        f"- Imported issue notes: {imported}\n"
        f"- Skipped low-signal rows: {skipped}\n"
        f"- Scope assumption: these support records were integrated as GreyOrange-related field issues because the dominant sites align with GreyOrange deployments.\n\n"
        f"## Site Distribution\n\n"
        f"{chr(10).join(f'- {site}: {count}' for site, count in per_site.most_common(15))}\n\n"
        f"## How To Use\n\n"
        f"- Treat each imported issue as historical support context, not as a pristine ticket.\n"
        f"- Link repeated patterns into site notes, SOPs, or product-quality reviews.\n"
        f"- Use [[09 Work/Issues/Issue Index|Issue Index]] and site pages to review current support surface."
    )
    write_note(VAULT / "99 System/Wellwit Issue Import Summary.md", summary_fm, summary_body)

    return {"issues": imported, "skipped": skipped}


def clean_obsolete():
    obsolete = [
        VAULT / "09 Work/Accounts/GreyOrange.md",
        VAULT / "09 Work/Accounts/Devonics.md",
    ]
    # These files are intentionally rewritten by the import functions above.
    return {"rewritten_account_notes": len(obsolete)}


def main():
    results = {
        "greyorange": import_greyorange(),
        "devonics": import_devonics(),
        "wellwit_issues": import_wellwit_issues(),
        "cleanup": clean_obsolete(),
    }
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
